import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side

st.set_page_config(page_title="MasterCmd Parameter Assignment Configurator", layout="wide")
st.title("⚙️ MasterCmd Parameter Assignment Configurator")

# --- Sidebar Configuration ---
st.sidebar.header("Configuration Settings")

num_blocks = st.sidebar.number_input("Blocks per Device", min_value=1, value=10, step=1)

nodes_sequence = st.sidebar.text_area(
    "Enter Node Numbers (comma-separated)",
    "26,27,28,29,30,31,32,33,34,35",
)
node_numbers = [int(n.strip()) for n in nodes_sequence.split(",") if n.strip().isdigit()]
num_devices = len(node_numbers)

entered_device_count = st.sidebar.number_input("Number of Devices (auto-calculated)", value=num_devices, step=1)
if entered_device_count != num_devices:
    st.sidebar.warning(f"⚠️ Entered number ({entered_device_count}) differs from detected nodes ({num_devices}).")

# --- Block Parameter Configuration ---
st.subheader("🧱 Block Parameter Configuration")

st.markdown("""
Enter configuration values for each block:
- Leave **Enable** blank → treated as 0  
- **Enable** may be 0, 1, or 2  
- If **Enable = 0**, you may leave **Func**, **DevAddress**, and **Count** empty  
""")

default_data = {
    "Block No.": list(range(1, num_blocks + 1)),
    "Enable": [1] * num_blocks,
    "Func": [4] * num_blocks,
    "DevAddress": [100 + i for i in range(num_blocks)],
    "Count": [1] * num_blocks,
}
block_df = pd.DataFrame(default_data)
edited_df = st.data_editor(block_df, num_rows="dynamic", key="block_table")

# --- Function Configuration ---
st.subheader("🧩 Function Configuration")

# Extract valid function IDs (non-empty numeric)
func_ids = sorted(set([int(f) for f in edited_df["Func"].dropna().unique() if str(f).isdigit()]))
if not func_ids:
    func_ids = [4]  # default if all blank

func_config = {}
cols = st.columns(len(func_ids))
for i, func_id in enumerate(func_ids):
    with cols[i]:
        st.markdown(f"**Function Code {func_id}**")
        init_addr = st.number_input(f"Initial Internal Address", min_value=0, value=0, key=f"init_{func_id}")
        offset = st.number_input(f"New Device Offset", min_value=0, value=10, key=f"offset_{func_id}")
        func_config[func_id] = {"initial": init_addr, "offset": offset}

# --- Action Buttons ---
col1, col2 = st.columns([1, 1])
preview_clicked = col1.button("👁️ Preview Output")
generate_clicked = col2.button("💾 Generate Excel")


# === Excel Generation Logic ===
def generate_excel():
    rows = []
    intaddress_track = {f: func_config[f]["initial"] for f in func_ids}

    for dev_idx, node in enumerate(node_numbers, start=1):
        for _, row in edited_df.iterrows():
            block_no = int(row["Block No."])
            enable = int(row["Enable"]) if not pd.isna(row["Enable"]) else 0

            # Skip Func, DevAddr, Count if Enable == 0 or blank
            func = int(row["Func"]) if not pd.isna(row["Func"]) and enable > 0 else None
            devaddr = int(row["DevAddress"]) if not pd.isna(row["DevAddress"]) and enable > 0 else None
            count = int(row["Count"]) if not pd.isna(row["Count"]) and enable > 0 else None

            # Compute IntAddress
            if enable > 0 and func in intaddress_track:
                if dev_idx > 1 and block_no == 1:
                    # Apply offset only for first block of that function on new device
                    intaddress_track[func] += func_config[func]["offset"]
                intaddr = intaddress_track[func]
                intaddress_track[func] += count if count else 0
            else:
                intaddr = ""

            params = [
                ("Enable", enable),
                ("Func", func if func is not None else ""),
                ("DevAddress", devaddr if devaddr is not None else ""),
                ("Count", count if count is not None else ""),
                ("IntAddress", intaddr),
                ("Node", node),
            ]
            for pname, val in params:
                rows.append({
                    "Device No.": dev_idx,
                    "Block No.": block_no,
                    "Node No.": node,
                    "Parameter": f"Cmd[{block_no}].{pname}",
                    "ConfigValue": val,
                })

    df = pd.DataFrame(rows)

    # Excel formatting
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Port1"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Add color bands and borders per device
    thin = Side(border_style="thick", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    fills = ["ADD8E6", "90EE90", "FFD580", "FFB6C1", "D3D3D3"]

    for idx, device in enumerate(df["Device No."].unique(), start=1):
        color = fills[(idx - 1) % len(fills)]
        dev_rows = df[df["Device No."] == device].index
        for i in dev_rows:
            for c in range(1, len(df.columns) + 1):
                ws.cell(row=i + 2, column=c).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, len(df.columns) + 1):
            ws.cell(row=dev_rows.min() + 2, column=c).border = border
            ws.cell(row=dev_rows.max() + 2, column=c).border = border

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return df, output


# --- Preview ---
if preview_clicked:
    preview_df, _ = generate_excel()
    st.dataframe(preview_df)

# --- Generate Excel ---
if generate_clicked:
    df, output = generate_excel()
    st.success("✅ Excel file generated successfully!")
    st.download_button("⬇️ Download Excel", data=output, file_name="MasterCmd_Sequence_R01.xlsx")


# --- Save/Load Config ---
st.sidebar.subheader("⚙️ Save / Load Configuration")
if st.sidebar.button("💾 Save Current Configuration"):
    config_data = {
        "blocks": edited_df.to_dict(),
        "func_config": func_config,
        "nodes": node_numbers,
    }
    st.sidebar.download_button(
        label="⬇️ Download Config JSON",
        data=pd.Series(config_data).to_json(),
        file_name="config.json",
    )

uploaded_config = st.sidebar.file_uploader("Upload Config JSON", type=["json"])
if uploaded_config:
    import json
    loaded = json.load(uploaded_config)
    st.session_state["block_table"] = pd.DataFrame(loaded["blocks"])
    st.sidebar.success("✅ Configuration loaded successfully!")
